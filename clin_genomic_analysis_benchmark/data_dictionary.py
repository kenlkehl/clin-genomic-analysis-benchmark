"""Parse BPC variable_synopsis Excel workbooks into structured JSON.

Prefers `simple_variable_synopsis.xlsx` (single sheet, 5 cols: Dataset, Variable Name,
Field Label, Data Type, Values) over the full `BPC_*_variable_synopsis.xlsx`.
"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Optional

import openpyxl

from .cohorts import Cohort, find_data_dictionary
from .config import CACHE_DIR


@dataclass
class Variable:
    dataset: str
    name: str
    label: str
    data_type: str
    values: list[str]   # parsed bullet list of allowed values / format hints

    def to_dict(self) -> dict:
        return asdict(self)


def _split_values(raw: Optional[str]) -> list[str]:
    """Split the bullet-list string in the 'Values' column into clean entries.

    The cell uses a `•\t` bullet prefix and `_x000D_\n` (CRLF artefacts from XLSX)
    as line separators.
    """
    if not raw:
        return []
    s = str(raw).replace("_x000D_", "").replace("\r", "")
    # Split on bullets or newlines
    parts: list[str] = []
    for chunk in s.split("\n"):
        chunk = chunk.strip()
        if chunk.startswith("•"):
            chunk = chunk[1:].strip()
        if chunk:
            parts.append(chunk)
    return parts


def parse(xlsx_path: Path) -> list[Variable]:
    """Parse a simple_variable_synopsis.xlsx workbook into a list of Variables."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out: list[Variable] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        # Tolerate a couple known header shapes
        cols = [str(h).strip() if h is not None else "" for h in header]
        try:
            i_dataset = cols.index("Dataset")
            i_name = cols.index("Variable Name")
            i_label = cols.index("Field Label")
            i_type = cols.index("Data Type")
            i_values = cols.index("Values")
        except ValueError:
            continue  # not the simple synopsis shape
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            out.append(
                Variable(
                    dataset=str(row[i_dataset] or "").strip(),
                    name=str(row[i_name] or "").strip(),
                    label=str(row[i_label] or "").strip(),
                    data_type=str(row[i_type] or "").strip(),
                    values=_split_values(row[i_values]),
                )
            )
    return out


def by_dataset(variables: list[Variable]) -> dict[str, list[Variable]]:
    out: dict[str, list[Variable]] = {}
    for v in variables:
        out.setdefault(v.dataset, []).append(v)
    return out


def to_compact_markdown(variables: list[Variable], max_value_chars: int = 200) -> str:
    """Render the dictionary as compact Markdown for inclusion in LLM prompts.

    One section per dataset; one line per variable; allowed values truncated.
    """
    grouped = by_dataset(variables)
    lines: list[str] = []
    for ds in sorted(grouped):
        lines.append(f"\n## {ds}\n")
        for v in grouped[ds]:
            vals = " | ".join(v.values)
            if len(vals) > max_value_chars:
                vals = vals[:max_value_chars].rstrip() + " …"
            lines.append(f"- **{v.name}** ({v.data_type}) — {v.label}" + (f"\n    - values: {vals}" if vals else ""))
    return "\n".join(lines)


def load(cohort: Cohort, use_cache: bool = True) -> list[Variable]:
    """Load and parse the data dictionary for a cohort, with on-disk caching."""
    cache = CACHE_DIR / "dictionary" / f"{cohort.name}.json"
    if use_cache and cache.exists():
        data = json.loads(cache.read_text())
        return [Variable(**d) for d in data]
    xlsx_path = find_data_dictionary(cohort)
    variables = parse(xlsx_path)
    cache.parent.mkdir(parents=True, exist_ok=True)
    cache.write_text(json.dumps([v.to_dict() for v in variables], indent=2))
    return variables
