#!/usr/bin/env python
"""Add canonical disambiguation concept IDs and the menu to the gold workbook.

The reviewed prose concepts remain untouched for human auditability.  This is a
one-time deterministic migration; future reviewers should edit the ID column
directly using the ``concept_menu`` sheet as the data dictionary.
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from clin_genomic_analysis_benchmark import config  # noqa: E402
from clin_genomic_analysis_benchmark.concepts import (  # noqa: E402
    CONCEPT_MENU,
    infer_legacy_concept_list,
    validate_concept_ids,
)


ID_COLUMN = "disambiguation_concept_ids"
MENU_SHEET = "concept_menu"


def migrate(path: Path, *, dry_run: bool = False, backup: bool = True) -> tuple[int, int]:
    wb = openpyxl.load_workbook(path)
    ws = wb["questions"]
    headers = [cell.value for cell in ws[1]]
    if "disambiguation_concepts" not in headers:
        raise ValueError("questions sheet has no disambiguation_concepts column")
    prose_col = headers.index("disambiguation_concepts") + 1
    class_col = headers.index("classification") + 1

    had_id_column = ID_COLUMN in headers
    if had_id_column:
        id_col = headers.index(ID_COLUMN) + 1
    else:
        id_col = prose_col + 1
        ws.insert_cols(id_col)
        ws.cell(1, id_col, ID_COLUMN)
        ws.cell(1, id_col).font = Font(bold=True)

    n_questions = 0
    n_ids = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, class_col).value != "ambiguous":
            ws.cell(row, id_col, None)
            continue
        prose = [
            value.strip()
            for value in str(ws.cell(row, prose_col).value or "").splitlines()
            if value.strip()
        ]
        existing_ids = [
            value.strip()
            for value in str(ws.cell(row, id_col).value or "").splitlines()
            if value.strip()
        ] if had_id_column else []
        ids = existing_ids or infer_legacy_concept_list(prose)
        errors = validate_concept_ids(ids)
        if errors:
            raise ValueError(f"row {row} has invalid concept IDs: {'; '.join(errors)}")
        if not ids:
            raise ValueError(f"row {row} is ambiguous but maps to no concept IDs")
        ws.cell(row, id_col, "\n".join(ids))
        ws.cell(row, id_col).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        n_questions += 1
        n_ids += len(ids)

    if MENU_SHEET in wb.sheetnames:
        del wb[MENU_SHEET]
    menu = wb.create_sheet(MENU_SHEET)
    menu.append(["concept_id", "label", "description"])
    for cell in menu[1]:
        cell.font = Font(bold=True)
    for concept in CONCEPT_MENU:
        menu.append([concept.id, concept.label, concept.description])
    menu.freeze_panes = "A2"
    menu.auto_filter.ref = menu.dimensions
    menu.column_dimensions["A"].width = 38
    menu.column_dimensions["B"].width = 34
    menu.column_dimensions["C"].width = 100

    legend = wb["legend"]
    legend_fields = [legend.cell(row, 1).value for row in range(1, legend.max_row + 1)]
    if ID_COLUMN not in legend_fields:
        prose_legend_row = legend_fields.index("disambiguation_concepts") + 1
        id_legend_row = prose_legend_row + 1
        legend.insert_rows(id_legend_row)
        legend.cell(id_legend_row, 1, ID_COLUMN)
        legend.cell(
            id_legend_row,
            2,
            "AMBIG: canonical IDs from concept_menu used for exact rules-based scoring. "
            "Newline-separated.",
        )

    if not dry_run:
        if backup:
            backup_path = path.with_name(path.stem + ".pre_rules_backup" + path.suffix)
            if not backup_path.exists():
                shutil.copy2(path, backup_path)
        wb.save(path)
    return n_questions, n_ids


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=config.workbook_path())
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()
    n_questions, n_ids = migrate(
        args.workbook,
        dry_run=args.dry_run,
        backup=not args.no_backup,
    )
    mode = "Validated" if args.dry_run else "Migrated"
    print(f"{mode} {n_questions} ambiguous questions with {n_ids} canonical concept IDs")
    print(f"Workbook: {args.workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
