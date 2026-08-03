#!/usr/bin/env python
"""Regenerate BOTH question banks from the current review workbook.

Source of truth: the review workbook under the gold root (CLINGEN_GOLD_ROOT).
Writes two banks:
  - GOLD  (out-of-repo, gold_root/questions/<cohort>.yaml): full, with answers —
    read only by the harness at scoring/compute time.
  - PUBLIC (in-repo, questions/<cohort>.yaml): gold-free (id/category/text only) —
    served to the agent under evaluation.

Maps the review columns into the CohortQuestionFile / Question Pydantic schema and validates.
Keeping the answers out of the in-repo bank is what prevents the agent from reading gold.
"""
from __future__ import annotations
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from clin_genomic_analysis_benchmark import config  # noqa: E402
from clin_genomic_analysis_benchmark.concepts import (  # noqa: E402
    infer_legacy_concept_list,
    validate_concept_ids,
)
ROOT = config.REPO_ROOT
from clin_genomic_analysis_benchmark.questions.schema import (  # noqa: E402
    CohortQuestionFile, Question, AnalysisSpec, GoldAnswer, SupportingEvidence,
)
from clin_genomic_analysis_benchmark.questions import io as q_io  # noqa: E402

SRC = config.workbook_path()
MODEL = "claude-opus-4-8@vertex"
DIAG_KEYS = {"cell_sizes", "cell_counts", "note", "notes", "small_cell_flag",
             "penalizer_used", "failure_reason"}

def parse_tables(gold_script: str) -> list[str]:
    # gold_script values are like "gold_standard/<cohort>/<qid>.py", now rooted
    # under the (out-of-repo) gold root.
    p = config.gold_root() / gold_script
    if not gold_script or not p.exists():
        return ["cancer_level_dataset_index.csv"]
    t = p.read_text()
    tabs = set(re.findall(r'"([A-Za-z0-9_./-]+\.(?:csv|txt))"', t)) | \
           set(re.findall(r"'([A-Za-z0-9_./-]+\.(?:csv|txt))'", t))
    tabs = sorted({Path(x).name for x in tabs})
    return tabs or ["cancer_level_dataset_index.csv"]

def population_unit(text: str, atype: str) -> str:
    tl = (text or "").lower()
    if re.search(r"\bsamples?\b", tl):
        return "sample"
    if "per index" in tl or "per colorectal" in tl or "per prostate" in tl or "per pancreatic" in tl:
        return "cancer"
    return "patient"

def build_gold_answer(qid: str, raw: str, atype: str) -> dict:
    full = json.loads(raw)
    d = {k: v for k, v in full.items() if k not in DIAG_KEYS}
    if qid == "prostate_1.2-Qf17acd7c":   # well-posed but not estimable in this cohort
        d["unanswerable"] = True
        d["unanswerable_reason"] = str(
            full.get("failure_reason") or full.get("value") or "estimand is not identifiable")
    return d


def parse_concept_ids(raw: object, prose_concepts: list[str]) -> list[str]:
    """Read canonical IDs, falling back to deterministic legacy migration."""
    if raw:
        text = str(raw).strip()
        try:
            decoded = json.loads(text)
        except json.JSONDecodeError:
            decoded = None
        if isinstance(decoded, list):
            ids = [str(value).strip() for value in decoded if str(value).strip()]
        else:
            ids = [value.strip() for value in re.split(r"[\n,]", text) if value.strip()]
    else:
        ids = infer_legacy_concept_list(prose_concepts)
    errors = validate_concept_ids(ids)
    if errors:
        raise ValueError("; ".join(errors))
    return ids

def main() -> int:
    wb = openpyxl.load_workbook(SRC)
    ws = wb["questions"]
    hdr = [c.value for c in ws[1]]
    ci = {n: i + 1 for i, n in enumerate(hdr)}
    def g(row: int, name: str):
        return ws.cell(row, ci[name]).value

    by_cohort: dict[str, list[Question]] = {}
    n_un = n_amb = 0
    for r in range(2, ws.max_row + 1):
        qid = g(r, "qid")
        cohort = g(r, "cohort")
        cls = g(r, "classification")
        common = dict(
            id=qid, category=int(g(r, "category")), text=g(r, "question_text"),
            classification=cls, source=(g(r, "source") or "llm"), review_status="reviewed",
        )
        if cls == "unambiguous":
            n_un += 1
            atype = g(r, "expected_answer_type")
            gold_script = g(r, "gold_script") or ""
            q = Question(
                **common,
                analysis_spec=AnalysisSpec(
                    population_unit=population_unit(g(r, "question_text"), atype),
                    tables=parse_tables(gold_script),
                    statistic=(g(r, "analysis_plan_summary") or f"{atype} per analysis plan").strip(),
                    expected_answer_type=atype,
                ),
                gold_answer=GoldAnswer(**build_gold_answer(qid, g(r, "gold_answer"), atype)),
                gold_supporting_evidence=SupportingEvidence(gold_script=gold_script or None),
            )
        else:
            n_amb += 1
            concepts = [ln.strip() for ln in str(g(r, "disambiguation_concepts") or "").split("\n") if ln.strip()]
            raw_ids = (g(r, "disambiguation_concept_ids")
                       if "disambiguation_concept_ids" in ci else None)
            try:
                concept_ids = parse_concept_ids(raw_ids, concepts)
            except ValueError as exc:
                raise ValueError(f"{qid}: invalid disambiguation concept IDs: {exc}") from exc
            q = Question(
                **common,
                disambiguation_concept_ids=concept_ids,
                disambiguation_concepts=concepts,
            )
        by_cohort.setdefault(cohort, []).append(q)

    now = datetime.now(timezone.utc)
    total = 0
    for cohort, qs in by_cohort.items():
        cqf = CohortQuestionFile(cohort=cohort, generated_at=now, model=MODEL, questions=qs)
        gold_path = q_io.save_gold(cqf)                       # full, out-of-repo
        pub_path = q_io.save_public(q_io.to_public(cqf))      # gold-free, in-repo
        total += len(qs)
        print(f"  {cohort}: {len(qs)} questions -> gold {gold_path}  |  public {pub_path}")
    print(f"TOTAL: {total} questions ({n_un} unambiguous / {n_amb} ambiguous) across {len(by_cohort)} cohorts")
    print(f"GOLD ROOT: {config.gold_root()}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
